#!/usr/bin/env python3
"""Local Marvel Rivals match database.

This intentionally avoids Tracker.gg scraping. It can import your existing
Tracker.gg-derived workbook and can update from an API source you are allowed
to use, such as MarvelRivalsAPI.com with your own key.
"""

from __future__ import annotations

import argparse
import csv
import email
import html
import re
import hashlib
import json
import os
import sqlite3
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from email import policy
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable


ROOT = Path(__file__).resolve().parent
DEFAULT_DB = ROOT / "marvel_rivals.sqlite3"
SCHEMA = ROOT / "schema.sql"


MATCH_COLUMNS = [
    "source",
    "source_match_id",
    "match_key",
    "match_date",
    "match_timestamp",
    "relative_time",
    "match_type",
    "game_type",
    "map",
    "hero",
    "result",
    "award",
    "score_for",
    "score_against",
    "rank",
    "rating_score",
    "rank_delta",
    "kills",
    "deaths",
    "assists",
    "kda_reported",
    "win_flag",
    "loss_flag",
    "draw_flag",
    "raw_json",
]

GAME_TYPES = {"Convergence", "Convoy", "Domination"}

RANK_WORDS = {
    "Bronze",
    "Silver",
    "Gold",
    "Platinum",
    "Diamond",
    "Grandmaster",
    "Celestial",
    "Eternity",
    "One Above All",
}

HEROES = {
    "Adam Warlock",
    "Black Panther",
    "Black Widow",
    "Bruce Banner",
    "Captain America",
    "Cloak & Dagger",
    "Doctor Strange",
    "Groot",
    "Hawkeye",
    "Hela",
    "Hulk",
    "Human Torch",
    "Invisible Woman",
    "Iron Fist",
    "Iron Man",
    "Jeff The Land Shark",
    "Loki",
    "Luna Snow",
    "Magik",
    "Magneto",
    "Mantis",
    "Mister Fantastic",
    "Moon Knight",
    "Namor",
    "Peni Parker",
    "Psylocke",
    "Punisher",
    "Rocket Raccoon",
    "Scarlet Witch",
    "Spider-Man",
    "Squirrel Girl",
    "Star-Lord",
    "Storm",
    "The Thing",
    "Thor",
    "Venom",
    "Winter Soldier",
    "Wolverine",
}

MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def connect(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db(db_path: Path) -> None:
    with connect(db_path) as conn:
        conn.executescript(SCHEMA.read_text(encoding="utf-8"))


def known_values(conn: sqlite3.Connection, column: str, fallback: set[str]) -> set[str]:
    rows = conn.execute(
        f"SELECT DISTINCT {column} AS value FROM matches WHERE {column} IS NOT NULL"
    ).fetchall()
    return {str(row["value"]) for row in rows if row["value"]} | fallback


def as_int(value: Any) -> int | None:
    if value in ("", None):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None


def as_float(value: Any) -> float | None:
    if value in ("", None):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def iso_date(value: Any) -> str | None:
    if value in ("", None):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text[:10]


def result_flags(result: str | None) -> tuple[int, int, int]:
    normalized = (result or "").strip().lower()
    return (
        1 if normalized == "win" else 0,
        1 if normalized == "loss" else 0,
        1 if "draw" in normalized or "unknown" in normalized else 0,
    )


def stable_key(parts: Iterable[Any]) -> str:
    text = "|".join("" if part is None else str(part) for part in parts)
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def tracker_natural_key(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        row.get("match_date"),
        row.get("match_type"),
        row.get("game_type"),
        row.get("map"),
        row.get("hero"),
        row.get("result"),
        row.get("score_for"),
        row.get("score_against"),
        row.get("rank"),
        row.get("rating_score"),
        row.get("rank_delta"),
        row.get("kills"),
        row.get("deaths"),
        row.get("assists"),
    )


def tracker_match_key(row: dict[str, Any]) -> str:
    return "tracker:" + stable_key(tracker_natural_key(row))


def find_tracker_duplicate(conn: sqlite3.Connection, row: dict[str, Any]) -> sqlite3.Row | None:
    if not str(row.get("source") or "").startswith("tracker_"):
        return None
    return conn.execute(
        """
        SELECT id, match_key
        FROM matches
        WHERE source LIKE 'tracker_%'
          AND COALESCE(match_date, '') = COALESCE(?, '')
          AND COALESCE(match_type, '') = COALESCE(?, '')
          AND COALESCE(game_type, '') = COALESCE(?, '')
          AND COALESCE(map, '') = COALESCE(?, '')
          AND COALESCE(hero, '') = COALESCE(?, '')
          AND COALESCE(result, '') = COALESCE(?, '')
          AND COALESCE(score_for, -999999) = COALESCE(?, -999999)
          AND COALESCE(score_against, -999999) = COALESCE(?, -999999)
          AND COALESCE(rank, '') = COALESCE(?, '')
          AND COALESCE(rating_score, -999999) = COALESCE(?, -999999)
          AND COALESCE(rank_delta, -999999) = COALESCE(?, -999999)
          AND COALESCE(kills, -999999) = COALESCE(?, -999999)
          AND COALESCE(deaths, -999999) = COALESCE(?, -999999)
          AND COALESCE(assists, -999999) = COALESCE(?, -999999)
        LIMIT 1
        """,
        tracker_natural_key(row),
    ).fetchone()


def upsert_match(conn: sqlite3.Connection, row: dict[str, Any]) -> str:
    natural_duplicate = find_tracker_duplicate(conn, row)
    if natural_duplicate:
        update_cols = [col for col in MATCH_COLUMNS if col not in {"match_key"}]
        assignments = ", ".join(f"{col} = ?" for col in update_cols)
        conn.execute(
            f"UPDATE matches SET {assignments}, updated_at = datetime('now') WHERE id = ?",
            [row.get(col) for col in update_cols] + [natural_duplicate["id"]],
        )
        return "updated"

    values = [row.get(col) for col in MATCH_COLUMNS]
    placeholders = ", ".join("?" for _ in MATCH_COLUMNS)
    updates = ", ".join(
        f"{col}=excluded.{col}"
        for col in MATCH_COLUMNS
        if col not in {"match_key", "created_at"}
    )
    exists = conn.execute(
        "SELECT 1 FROM matches WHERE match_key = ?",
        (row["match_key"],),
    ).fetchone()
    conn.execute(
        f"""
        INSERT INTO matches ({", ".join(MATCH_COLUMNS)})
        VALUES ({placeholders})
        ON CONFLICT(match_key) DO UPDATE SET
            {updates},
            updated_at=datetime('now')
        """,
        values,
    )
    return "updated" if exists else "inserted"


def record_import(
    conn: sqlite3.Connection,
    source: str,
    source_path: str | None,
    seen: int,
    inserted: int,
    updated: int,
    notes: str | None = None,
) -> None:
    conn.execute(
        """
        INSERT INTO imports
            (source, source_path, rows_seen, rows_inserted, rows_updated, notes)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (source, source_path, seen, inserted, updated, notes),
    )


def import_excel(db_path: Path, excel_path: Path) -> tuple[int, int, int]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit(
            "Missing dependency: openpyxl. Run `python3 -m pip install -r requirements.txt`."
        ) from exc

    init_db(db_path)
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "Competitive_Matches" not in workbook.sheetnames:
        raise SystemExit("Workbook does not contain a Competitive_Matches sheet.")

    sheet = workbook["Competitive_Matches"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(cell) if cell is not None else "" for cell in next(rows)]
    seen = inserted = updated = 0

    with connect(db_path) as conn:
        for raw_values in rows:
            raw = dict(zip(headers, raw_values))
            if not any(value is not None for value in raw.values()):
                continue
            seen += 1
            result = raw.get("Result")
            win, loss, draw = result_flags(result)
            row = {
                "source": "tracker_export_xlsx",
                "source_match_id": str(raw.get("Match_ID") or ""),
                "match_date": iso_date(raw.get("Match_Date")),
                "match_timestamp": None,
                "relative_time": raw.get("Relative_Time"),
                "match_type": raw.get("Match_Type"),
                "game_type": raw.get("Game_Type"),
                "map": raw.get("Map"),
                "hero": raw.get("Hero"),
                "result": result,
                "award": raw.get("Award"),
                "score_for": as_int(raw.get("Score_For")),
                "score_against": as_int(raw.get("Score_Against")),
                "rank": raw.get("Rank"),
                "rating_score": as_int(raw.get("Rating_Score")),
                "rank_delta": as_int(raw.get("Rank_Delta")),
                "kills": as_int(raw.get("Kills")),
                "deaths": as_int(raw.get("Deaths")),
                "assists": as_int(raw.get("Assists")),
                "kda_reported": as_float(raw.get("KDA_Reported")),
                "win_flag": win,
                "loss_flag": loss,
                "draw_flag": draw,
                "raw_json": json.dumps(raw, default=str, ensure_ascii=True),
            }
            row["match_key"] = tracker_match_key(row)
            status = upsert_match(conn, row)
            if status == "inserted":
                inserted += 1
            else:
                updated += 1
        record_import(conn, "tracker_export_xlsx", str(excel_path), seen, inserted, updated)

    print(f"Imported {seen} rows into {db_path} ({inserted} inserted, {updated} updated).")
    return seen, inserted, updated


class VisibleTextParser(HTMLParser):
    block_tags = {
        "article",
        "br",
        "div",
        "h1",
        "h2",
        "h3",
        "h4",
        "li",
        "p",
        "section",
        "span",
        "td",
        "th",
        "tr",
    }

    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []
        self.skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in {"script", "style", "noscript", "svg"}:
            self.skip_depth += 1
            return
        if self.skip_depth:
            return
        if tag in self.block_tags:
            self.parts.append("\n")
        for name, value in attrs:
            if name in {"alt", "aria-label", "title"} and value:
                self.parts.append(f"\n{value}\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style", "noscript", "svg"} and self.skip_depth:
            self.skip_depth -= 1
            return
        if not self.skip_depth and tag in self.block_tags:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if not self.skip_depth:
            self.parts.append(data)

    def text(self) -> str:
        return html.unescape(" ".join(self.parts))


def strip_html(raw: str) -> str:
    parser = VisibleTextParser()
    parser.feed(raw)
    return parser.text()


def clean_lines(text: str) -> list[str]:
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t\r\f\v]+", " ", text)
    lines = [line.strip() for line in text.splitlines()]
    return [line for line in lines if line]


def read_saved_file_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".mhtml", ".mht"}:
        message = email.message_from_binary_file(path.open("rb"), policy=policy.default)
        chunks: list[str] = []
        for part in message.walk():
            ctype = part.get_content_type()
            if ctype in {"text/html", "text/plain"}:
                payload = part.get_payload(decode=True)
                if payload:
                    charset = part.get_content_charset() or "utf-8"
                    chunks.append(payload.decode(charset, errors="replace"))
        return "\n".join(strip_html(chunk) if "<" in chunk and ">" in chunk else chunk for chunk in chunks)

    if suffix == ".pdf":
        try:
            import pypdf
        except ImportError as exc:
            raise SystemExit("Missing dependency: pypdf. Run `python3 -m pip install -r requirements.txt`.") from exc
        reader = pypdf.PdfReader(str(path))
        return "\n".join(page.extract_text() or "" for page in reader.pages)

    raw = path.read_text(encoding="utf-8", errors="replace")
    if suffix in {".html", ".htm"} or "<html" in raw.lower():
        return strip_html(raw)
    return raw


def infer_snapshot_date(path: Path) -> datetime:
    for candidate in [path.stem, path.name]:
        match = re.search(r"(20\d{2})[-_ ]?(\d{2})[-_ ]?(\d{2})", candidate)
        if match:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return datetime.fromtimestamp(path.stat().st_mtime)


def parse_date_label(label: str, snapshot: datetime) -> str | None:
    text = label.strip().lower().rstrip(",")
    if text == "today":
        return snapshot.date().isoformat()
    if text == "yesterday":
        return datetime.fromtimestamp(snapshot.timestamp() - 86400).date().isoformat()
    match = re.match(r"([a-z]+)\.?\s+(\d{1,2})(?:,\s*(\d{4}))?$", text)
    if not match:
        return None
    month = MONTHS.get(match.group(1))
    if not month:
        return None
    day = int(match.group(2))
    year = int(match.group(3) or snapshot.year)
    parsed = datetime(year, month, day)
    if not match.group(3) and parsed.date() > snapshot.date():
        parsed = datetime(year - 1, month, day)
    return parsed.date().isoformat()


def find_known_value(block: list[str], known: set[str]) -> str | None:
    joined = "\n".join(block)
    for value in sorted(known, key=len, reverse=True):
        if re.search(rf"(?<!\w){re.escape(value)}(?!\w)", joined, flags=re.IGNORECASE):
            return value
    return None


def number_after_label(block: list[str], labels: set[str]) -> int | None:
    for index, line in enumerate(block):
        normalized = line.strip().lower()
        if normalized in labels:
            for next_line in block[index + 1 : index + 5]:
                value = as_int(re.sub(r"[^\d.-]", "", next_line))
                if value is not None:
                    return value
        for label in labels:
            match = re.search(rf"\b{re.escape(label)}\b\s*[:\-]?\s*(-?\d+)", normalized)
            if match:
                return as_int(match.group(1))
    return None


def parse_score(block_text: str) -> tuple[int | None, int | None]:
    patterns = [
        r"\bscore\b\s*[:\-]?\s*(\d+)\s*[-:]\s*(\d+)",
        r"\b(\d+)\s*[-:]\s*(\d+)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, block_text, flags=re.IGNORECASE)
        if match:
            return as_int(match.group(1)), as_int(match.group(2))
    return None, None


def parse_rank(block: list[str]) -> str | None:
    joined = "\n".join(block)
    for rank in sorted(RANK_WORDS, key=len, reverse=True):
        match = re.search(rf"\b{re.escape(rank)}(?:\s+[IVX]+)?\b", joined, flags=re.IGNORECASE)
        if match:
            return match.group(0).title()
    return None


def parse_rating_and_delta(block: list[str]) -> tuple[int | None, int | None]:
    joined = " ".join(block)
    rating = None
    for match in re.finditer(r"\b([1-9]\d{3,4})\b", joined):
        value = as_int(match.group(1))
        if value and 1000 <= value <= 99999:
            rating = value
            break
    delta = number_after_label(block, {"rank delta", "rating delta", "delta"})
    if delta is None:
        match = re.search(r"(?<!\d)([+-]\d{1,3})(?!\d)", joined)
        if match:
            delta = as_int(match.group(1))
    return rating, delta


def parse_saved_tracker_rows(
    path: Path,
    text: str,
    known_heroes: set[str],
    known_maps: set[str],
) -> list[dict[str, Any]]:
    lines = clean_lines(text)
    snapshot = infer_snapshot_date(path)
    rows: list[dict[str, Any]] = []
    competitive_indexes = [
        index for index, line in enumerate(lines)
        if line.strip().lower() == "competitive"
    ]

    for position, index in enumerate(competitive_indexes):
        end = competitive_indexes[position + 1] if position + 1 < len(competitive_indexes) else min(len(lines), index + 80)
        context_start = max(0, index - 16)
        block = lines[context_start:end]
        future_block = lines[index:end]
        block_text = "\n".join(block)

        result = None
        if re.search(r"\bwin\b", block_text, flags=re.IGNORECASE):
            result = "Win"
        if re.search(r"\bloss\b|\blost\b", block_text, flags=re.IGNORECASE):
            result = "Loss"
        if re.search(r"\bdraw\b|\bunknown\b", block_text, flags=re.IGNORECASE):
            result = "Draw/Unknown"

        hero = find_known_value(future_block, known_heroes)
        game_type = find_known_value(future_block, GAME_TYPES)
        map_name = find_known_value(future_block, known_maps)
        if not hero or not result:
            continue

        match_date = None
        for previous in reversed(lines[max(0, index - 20) : index + 1]):
            match_date = parse_date_label(previous, snapshot)
            if match_date:
                break
        if not match_date:
            relative = re.search(r"\b(\d+)\s*h(?:ours?)?\s+ago\b", block_text, flags=re.IGNORECASE)
            if relative and as_int(relative.group(1)) is not None and as_int(relative.group(1)) < 24:
                match_date = snapshot.date().isoformat()

        score_for, score_against = parse_score(block_text)
        rating_score, rank_delta = parse_rating_and_delta(future_block)
        kills = number_after_label(future_block, {"kills", "k"})
        deaths = number_after_label(future_block, {"deaths", "d"})
        assists = number_after_label(future_block, {"assists", "a"})
        kda = number_after_label(future_block, {"kda"})
        if kills is None or deaths is None or assists is None:
            numeric_lines = [as_int(line) for line in future_block if re.fullmatch(r"-?\d+", line)]
            numeric_lines = [value for value in numeric_lines if value is not None]
            if len(numeric_lines) >= 3:
                kills = kills if kills is not None else numeric_lines[-3]
                deaths = deaths if deaths is not None else numeric_lines[-2]
                assists = assists if assists is not None else numeric_lines[-1]

        win, loss, draw = result_flags(result)
        row = {
            "source": "tracker_saved_file",
            "source_match_id": None,
            "match_date": match_date,
            "match_timestamp": None,
            "relative_time": next((line for line in block if re.search(r"\bago\b", line, re.I)), None),
            "match_type": "Competitive",
            "game_type": game_type,
            "map": map_name,
            "hero": hero,
            "result": result,
            "award": "MVP" if re.search(r"\bMVP\b", block_text) else ("SVP" if re.search(r"\bSVP\b", block_text) else None),
            "score_for": score_for,
            "score_against": score_against,
            "rank": parse_rank(future_block),
            "rating_score": rating_score,
            "rank_delta": rank_delta,
            "kills": kills,
            "deaths": deaths,
            "assists": assists,
            "kda_reported": as_float(kda),
            "win_flag": win,
            "loss_flag": loss,
            "draw_flag": draw,
            "raw_json": json.dumps({"source_file": str(path), "block": block}, ensure_ascii=True),
        }
        row["match_key"] = tracker_match_key(row)
        rows.append(row)

    return rows


def import_saved_file(db_path: Path, path: Path) -> tuple[int, int, int]:
    init_db(db_path)
    text = read_saved_file_text(path)
    with connect(db_path) as conn:
        known_heroes = known_values(conn, "hero", HEROES)
        known_maps = known_values(conn, "map", set())
        rows = parse_saved_tracker_rows(path, text, known_heroes, known_maps)
        seen = inserted = updated = 0
        for row in rows:
            seen += 1
            status = upsert_match(conn, row)
            if status == "inserted":
                inserted += 1
            else:
                updated += 1
        record_import(
            conn,
            "tracker_saved_file",
            str(path),
            seen,
            inserted,
            updated,
            notes="Parsed from saved MHTML/HTML/PDF/text export.",
        )
    return seen, inserted, updated


def import_folder(db_path: Path, folder: Path) -> None:
    supported = {".xlsx", ".xlsm", ".mhtml", ".mht", ".html", ".htm", ".pdf", ".txt"}
    if not folder.exists():
        folder.mkdir(parents=True)
        print(f"Created imports folder: {folder}")
        return

    files = [
        path for path in sorted(folder.iterdir())
        if path.is_file() and path.suffix.lower() in supported
    ]
    if not files:
        print(f"No supported import files found in {folder}")
        return

    total_seen = total_inserted = total_updated = 0
    for path in files:
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            seen, inserted, updated = import_excel(db_path, path)
            total_seen += seen
            total_inserted += inserted
            total_updated += updated
            continue
        seen, inserted, updated = import_saved_file(db_path, path)
        total_seen += seen
        total_inserted += inserted
        total_updated += updated
        print(f"{path.name}: {seen} parsed ({inserted} inserted, {updated} already present/updated).")

    print(f"Folder import complete: {total_seen} parsed ({total_inserted} inserted, {total_updated} updated).")


def fetch_json(url: str, api_key: str) -> dict[str, Any]:
    request = urllib.request.Request(url, headers={"x-api-key": api_key})
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise SystemExit(f"API request failed: HTTP {exc.code} {body[:500]}") from exc


def normalize_api_match(item: dict[str, Any]) -> dict[str, Any]:
    player = item.get("match_player") or {}
    hero = player.get("player_hero") or {}
    score = player.get("score_info") or {}
    is_win = player.get("is_win") or {}
    timestamp = as_int(item.get("match_time_stamp"))
    match_date = (
        datetime.fromtimestamp(timestamp, tz=timezone.utc).date().isoformat()
        if timestamp
        else None
    )
    won = bool(is_win.get("is_win"))
    result = "Win" if won else "Loss"
    win, loss, draw = result_flags(result)
    source_match_id = str(item.get("match_uid") or "")
    rank_delta = as_int(score.get("add_score"))

    row = {
        "source": "marvelrivalsapi",
        "source_match_id": source_match_id,
        "match_date": match_date,
        "match_timestamp": timestamp,
        "relative_time": None,
        "match_type": "Competitive",
        "game_type": str(item.get("game_mode_id") or item.get("play_mode_id") or ""),
        "map": str(item.get("match_map_id") or item.get("map_thumbnail") or ""),
        "hero": hero.get("hero_name"),
        "result": result,
        "award": None,
        "score_for": as_int(is_win.get("score")),
        "score_against": None,
        "rank": None,
        "rating_score": as_int(score.get("new_score")),
        "rank_delta": rank_delta,
        "kills": as_int(player.get("kills") if player.get("kills") is not None else hero.get("kills")),
        "deaths": as_int(player.get("deaths") if player.get("deaths") is not None else hero.get("deaths")),
        "assists": as_int(player.get("assists") if player.get("assists") is not None else hero.get("assists")),
        "kda_reported": None,
        "win_flag": win,
        "loss_flag": loss,
        "draw_flag": draw,
        "raw_json": json.dumps(item, default=str, ensure_ascii=True),
    }
    row["match_key"] = (
        f"marvelrivalsapi:{source_match_id}"
        if source_match_id
        else stable_key(
            [
                row["source"],
                row["match_timestamp"],
                row["game_type"],
                row["map"],
                row["hero"],
                row["result"],
                row["kills"],
                row["deaths"],
                row["assists"],
            ]
        )
    )
    return row


def update_api(
    db_path: Path,
    player: str,
    api_key: str,
    season: str | None,
    game_mode: str | None,
    limit: int,
    max_pages: int,
) -> None:
    init_db(db_path)
    seen = inserted = updated = 0

    with connect(db_path) as conn:
        for page in range(1, max_pages + 1):
            query = {
                "page": page,
                "limit": limit,
            }
            if season:
                query["season"] = season
            if game_mode:
                query["game_mode"] = game_mode
            url = (
                "https://marvelrivalsapi.com/api/v2/player/"
                + urllib.parse.quote(player)
                + "/match-history?"
                + urllib.parse.urlencode(query)
            )
            payload = fetch_json(url, api_key)
            history = payload.get("match_history") or []
            if not history:
                break
            for item in history:
                seen += 1
                status = upsert_match(conn, normalize_api_match(item))
                if status == "inserted":
                    inserted += 1
                else:
                    updated += 1
            pagination = payload.get("pagination") or {}
            if not pagination.get("has_more"):
                break
            time.sleep(0.25)
        record_import(
            conn,
            "marvelrivalsapi",
            f"player={player}",
            seen,
            inserted,
            updated,
            notes=f"season={season or ''}; game_mode={game_mode or ''}",
        )

    print(f"Fetched {seen} API rows into {db_path} ({inserted} inserted, {updated} updated).")


def report(db_path: Path, export_csv: Path | None = None) -> None:
    init_db(db_path)
    queries = {
        "Overall": """
            SELECT COUNT(*) AS matches, SUM(win_flag) AS wins, SUM(loss_flag) AS losses,
                   SUM(draw_flag) AS draws,
                   ROUND(SUM(win_flag) * 100.0 / NULLIF(COUNT(*), 0), 2) AS win_rate_pct,
                   SUM(rank_delta) AS total_rank_delta
            FROM matches
        """,
        "Top heroes": """
            SELECT hero, matches, wins, losses, win_rate, avg_kda, avg_rank_delta, total_rank_delta
            FROM hero_summary
            ORDER BY matches DESC
            LIMIT 10
        """,
        "Recent days": """
            SELECT match_date, matches, wins, losses, draws, win_rate, total_rank_delta
            FROM daily_summary
            ORDER BY match_date DESC
            LIMIT 14
        """,
    }

    with connect(db_path) as conn:
        for title, sql in queries.items():
            print(f"\n{title}")
            rows = conn.execute(sql).fetchall()
            for row in rows:
                print(dict(row))

        if export_csv:
            rows = conn.execute("SELECT * FROM matches ORDER BY match_date DESC, id DESC").fetchall()
            with export_csv.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=rows[0].keys() if rows else ["id"])
                writer.writeheader()
                for row in rows:
                    writer.writerow(dict(row))
            print(f"\nExported matches to {export_csv}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build and update a Marvel Rivals SQLite database.")
    parser.add_argument("--db", type=Path, default=DEFAULT_DB, help=f"SQLite DB path. Default: {DEFAULT_DB}")
    sub = parser.add_subparsers(dest="command", required=True)

    sub.add_parser("init", help="Create the database schema.")

    import_parser = sub.add_parser("import-excel", help="Import the existing Excel workbook.")
    import_parser.add_argument("excel_path", type=Path)

    folder_parser = sub.add_parser("import-folder", help="Import saved Tracker exports from a folder.")
    folder_parser.add_argument(
        "folder",
        type=Path,
        nargs="?",
        default=ROOT / "imports",
        help="Folder containing .xlsx, .mhtml, .html, .pdf, or .txt exports. Default: ./imports",
    )

    api_parser = sub.add_parser("update-api", help="Update from MarvelRivalsAPI.com.")
    api_parser.add_argument("--player", default=os.getenv("MARVEL_RIVALS_PLAYER"), required=not os.getenv("MARVEL_RIVALS_PLAYER"))
    api_parser.add_argument("--api-key", default=os.getenv("MARVEL_RIVALS_API_KEY"), required=not os.getenv("MARVEL_RIVALS_API_KEY"))
    api_parser.add_argument("--season", default=os.getenv("MARVEL_RIVALS_SEASON"))
    api_parser.add_argument("--game-mode", default=os.getenv("MARVEL_RIVALS_GAME_MODE"))
    api_parser.add_argument("--limit", type=int, default=int(os.getenv("MARVEL_RIVALS_LIMIT", "50")))
    api_parser.add_argument("--max-pages", type=int, default=int(os.getenv("MARVEL_RIVALS_MAX_PAGES", "2")))

    report_parser = sub.add_parser("report", help="Print summary tables.")
    report_parser.add_argument("--export-csv", type=Path)

    args = parser.parse_args()
    if args.command == "init":
        init_db(args.db)
        print(f"Initialized {args.db}")
    elif args.command == "import-excel":
        import_excel(args.db, args.excel_path)
    elif args.command == "import-folder":
        import_folder(args.db, args.folder)
    elif args.command == "update-api":
        update_api(args.db, args.player, args.api_key, args.season, args.game_mode, args.limit, args.max_pages)
    elif args.command == "report":
        report(args.db, args.export_csv)


if __name__ == "__main__":
    main()
